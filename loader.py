"""
CSV / Excel loader for project schedule data.

Expected CSV columns (MS-Project export style):
  row_num, WBS, Task Name, Duration, Start, Finish, Predecessors, % Complete

The loader normalises dates, infers WBS depth, flags summary rows,
then returns a list of dicts ready for db.upsert_tasks().
"""
import re
import ast
import pandas as pd
from io import StringIO, BytesIO
from typing import Union
from pathlib import Path
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

# ─── Date parsing ─────────────────────────────────────────────────────────────
_DATE_FMTS = [
    "%a %m/%d/%y %I:%M %p",   # Mon 10/16/17 8:00 AM
    "%m/%d/%y",
    "%m/%d/%Y",
    "%Y-%m-%d",
]


def _parse_date(raw) -> str | None:
    if pd.isna(raw) or str(raw).strip() == "":
        return None
    s = str(raw).strip()
    for fmt in _DATE_FMTS:
        try:
            return pd.to_datetime(s, format=fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return pd.to_datetime(s, infer_datetime_format=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def _parse_float(raw) -> float | None:
    if pd.isna(raw) or str(raw).strip() == "":
        return None
    s = str(raw).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


# ─── Duration parsing ─────────────────────────────────────────────────────────
def _parse_duration(raw) -> float | None:
    if pd.isna(raw) or str(raw).strip() == "":
        return None
    s = str(raw).strip().lower().replace("?", "")
    m = re.match(r"([\d.]+)\s*(day|days|d|hr|hrs|hour|hours|h|wk|week|weeks|w)?", s)
    if not m:
        return None
    val = float(m.group(1))
    unit = (m.group(2) or "day").lower()
    if unit.startswith("h"):
        val /= 8
    elif unit.startswith("w"):
        val *= 5
    return round(val, 2)


# ─── WBS helpers ─────────────────────────────────────────────────────────────
def _wbs_level(wbs: str) -> int:
    if not wbs or str(wbs).strip() == "":
        return 1
    return len(str(wbs).strip().split("."))


def _is_summary(wbs: str, task_name: str, duration: float | None) -> int:
    """Heuristic: a row is a summary if its WBS has <= 2 components OR duration is 0."""
    level = _wbs_level(wbs)
    if level <= 2:
        return 1
    if duration is not None and duration <= 0:
        return 1
    return 0


# ─── Main parse function ──────────────────────────────────────────────────────
def parse_schedule_csv(
    source: Union[str, Path, BytesIO, StringIO],
    project_id: int,
) -> list[dict]:
    """
    Parse a project schedule CSV and return a list of task dicts.

    Accepts a file path, a pathlib.Path, or a file-like object (BytesIO / StringIO).
    """
    if isinstance(source, (str, Path)):
        raw = pd.read_csv(source, header=None, dtype=str)
    else:
        raw = pd.read_csv(source, header=None, dtype=str)

    # The CSV has a leading empty column; find the header row
    # Strategy: first non-empty row that contains "Task Name"
    header_row_idx = None
    for i, row in raw.iterrows():
        if any("task name" in str(v).lower() for v in row.values):
            header_row_idx = i
            break

    if header_row_idx is None:
        # Treat first row as header
        header_row_idx = 0

    # Seek back to start for file-like objects (StringIO/BytesIO cursor is at EOF after first read)
    if hasattr(source, "seek"):
        source.seek(0)

    df = pd.read_csv(
        source,
        skiprows=header_row_idx if header_row_idx > 0 else None,
        dtype=str,
    )

    # Normalise column names
    df.columns = [str(c).strip().lower().replace(" ", "_").replace("%", "pct") for c in df.columns]

    # Map known column aliases
    _aliases = {
        "task_name": ["task_name", "name", "task"],
        "wbs": ["wbs"],
        "duration": ["duration", "dur"],
        "start": ["start", "start_date"],
        "finish": ["finish", "finish_date", "end", "end_date"],
        "predecessors": ["predecessors", "predecessor", "pred"],
        "pct_complete": ["pct_complete", "_%_complete", "complete", "done"],
    }

    col_map = {}
    for canonical, aliases in _aliases.items():
        for alias in aliases:
            if alias in df.columns:
                col_map[alias] = canonical
                break

    df = df.rename(columns=col_map)

    required = ["task_name"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"Could not find required column '{c}' in CSV. Columns found: {list(df.columns)}")

    # Fill missing optional columns
    for col, default in [("wbs", ""), ("duration", None), ("start", None),
                         ("finish", None), ("predecessors", ""), ("pct_complete", "0")]:
        if col not in df.columns:
            df[col] = default

    tasks = []
    for idx, row in df.iterrows():
        task_name = str(row.get("task_name", "")).strip()
        if not task_name or task_name.lower() in ("nan", "task name", ""):
            continue

        wbs = str(row.get("wbs", "")).strip()
        dur = _parse_duration(row.get("duration"))
        start = _parse_date(row.get("start"))
        finish = _parse_date(row.get("finish"))
        preds = str(row.get("predecessors", "")).strip()
        pct_raw = str(row.get("pct_complete", "0")).strip().replace("%", "").replace("?", "")
        try:
            pct = float(pct_raw) if pct_raw not in ("", "nan") else 0.0
        except ValueError:
            pct = 0.0

        # row_num: use the leading integer column if present (first col)
        row_num_val = None
        first_col = df.columns[0] if len(df.columns) > 0 else None
        if first_col and first_col not in col_map.values():
            try:
                row_num_val = int(float(str(row[first_col]).strip()))
            except (ValueError, TypeError):
                row_num_val = idx + 1
        else:
            row_num_val = idx + 1

        tasks.append(
            {
                "project_id": project_id,
                "row_num": row_num_val,
                "wbs": wbs if wbs != "nan" else "",
                "task_name": task_name,
                "duration_days": dur,
                "start_date": start,
                "finish_date": finish,
                "predecessors": preds if preds != "nan" else "",
                "pct_complete": pct,
                "wbs_level": _wbs_level(wbs),
                "is_summary": _is_summary(wbs, task_name, dur),
                "notes": "",
                "site_id": None,
                "template_id": None,
                "quantity": None,
                "unit": None,
            }
        )

    return tasks


def parse_schedule_excel(file_obj: BytesIO, project_id: int, sheet: int | str = 0) -> list[dict]:
    """Parse an Excel workbook (same column structure as CSV)."""
    df = pd.read_excel(file_obj, sheet_name=sheet, dtype=str)
    # Write to an in-memory CSV buffer and reuse csv parser
    buf = StringIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return parse_schedule_csv(buf, project_id)


def is_site_quantity_csv(source: Union[str, Path, BytesIO, StringIO]) -> bool:
    """Heuristic check for the Dakansy-style site schedule format."""
    sample = pd.read_csv(source, header=None, dtype=str, nrows=30)
    if hasattr(source, "seek"):
        source.seek(0)

    for _, row in sample.iterrows():
        values = [str(v).strip().lower() for v in row.values if str(v).strip() not in ("", "nan")]
        line = " ".join(values)
        if "activity" in line and "quantity" in line and "planned" in line:
            return True
    return False


def parse_site_quantity_csv(
    source: Union[str, Path, BytesIO, StringIO],
    project_id: int,
) -> list[dict]:
    """
    Parse site-quantity schedule CSV rows.

    Expected columns resemble:
      S.NO., ACTIVITY, QUANTITY, UNIT, PLANNED START, PLANNED DURATION,
      PLANNED FINISH, ACTUAL START, ACTUAL DURATION, ACTUAL FINISH, PERCENT COMPLETE
    """
    raw = pd.read_csv(source, header=None, dtype=str)

    header_row_idx = None
    for i, row in raw.iterrows():
        values = [str(v).strip().lower() for v in row.values if str(v).strip() not in ("", "nan")]
        line = " ".join(values)
        if "activity" in line and "quantity" in line and ("planned" in line or "percent complete" in line):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not locate site schedule header row.")

    if hasattr(source, "seek"):
        source.seek(0)

    df = pd.read_csv(source, skiprows=header_row_idx, dtype=str)
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    aliases = {
        "row_num": ["s.no.", "s.no", "s_no", "sr_no", "serial_no", "sl_no"],
        "task_name": ["activity", "task_name", "task"],
        "quantity": ["quantity", "qty"],
        "unit": ["unit", "uom"],
        "start": ["planned_start", "plan_start", "start"],
        "duration": ["planned_duration", "plan_duration", "duration"],
        "finish": ["planned_finish", "plan_finish", "finish"],
        "pct_complete": ["percent_complete", "%_complete", "pct_complete", "progress"],
    }

    mapped: dict[str, str] = {}
    for canonical, candidates in aliases.items():
        for c in candidates:
            if c in df.columns:
                mapped[canonical] = c
                break

    if "task_name" not in mapped:
        raise ValueError("Could not find ACTIVITY / task name column in site schedule.")

    tasks: list[dict] = []
    for idx, row in df.iterrows():
        task_name = str(row.get(mapped["task_name"], "")).strip()
        if not task_name or task_name.lower() in ("nan", "activity"):
            continue

        qty = _parse_float(row.get(mapped.get("quantity", ""))) if "quantity" in mapped else None
        unit = str(row.get(mapped.get("unit", ""), "")).strip() if "unit" in mapped else ""
        dur = _parse_duration(row.get(mapped.get("duration", ""))) if "duration" in mapped else None
        start = _parse_date(row.get(mapped.get("start", ""))) if "start" in mapped else None
        finish = _parse_date(row.get(mapped.get("finish", ""))) if "finish" in mapped else None

        pct = 0.0
        if "pct_complete" in mapped:
            pct_raw = str(row.get(mapped["pct_complete"], "0")).strip().replace("%", "")
            try:
                pct = float(pct_raw) if pct_raw not in ("", "nan") else 0.0
            except ValueError:
                pct = 0.0

        row_num = idx + 1
        if "row_num" in mapped:
            try:
                row_num = int(float(str(row.get(mapped["row_num"], "")).strip()))
            except (TypeError, ValueError):
                row_num = idx + 1

        # Section rows in this format usually have no quantity and no dates.
        is_summary = 1 if (qty is None and not start and not finish) else 0

        tasks.append(
            {
                "project_id": project_id,
                "row_num": row_num,
                "wbs": str(row_num),
                "task_name": task_name,
                "duration_days": dur,
                "start_date": start,
                "finish_date": finish,
                "predecessors": "",
                "pct_complete": pct,
                "wbs_level": 1,
                "is_summary": is_summary,
                "notes": "",
                "site_id": None,
                "template_id": None,
                "quantity": qty,
                "unit": unit if unit.lower() not in ("nan", "") else None,
            }
        )

    return tasks


def _col_letter_to_index(letter: str) -> int:
    out = 0
    for ch in letter:
        if not ch.isalpha():
            break
        out = out * 26 + (ord(ch.upper()) - ord("A") + 1)
    return out


def _safe_eval_numeric(expr: str) -> float | None:
    """Evaluate a simple numeric expression containing +, -, numbers, parentheses, and max()."""
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError:
        return None

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub)):
            left = _eval(node.left)
            right = _eval(node.right)
            if left is None or right is None:
                return None
            return left + right if isinstance(node.op, ast.Add) else left - right
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            val = _eval(node.operand)
            if val is None:
                return None
            return val if isinstance(node.op, ast.UAdd) else -val
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id == "max":
            args = [_eval(a) for a in node.args]
            if any(a is None for a in args):
                return None
            return max(args)
        return None

    return _eval(tree)


def _normalize_excel_formula(formula: str) -> str:
    return formula.strip().upper().replace("$", "").replace(" ", "")


def _extract_dependency_terms(formula: str) -> list[tuple[str, int, float]]:
    """
    Return dependency terms as tuples: (dep_type, from_row_num, lag_days).

    Mapping:
      E<row> refs -> SS dependency
      G<row> refs -> FS dependency
      lag from +/- integer applied directly to that reference token
    """
    norm = _normalize_excel_formula(formula)
    if not norm.startswith("="):
        return []

    terms: list[tuple[str, int, float]] = []
    for m in re.finditer(r"([EG])(\d+)([+-]\d+(?:\.\d+)?)?", norm):
        col = m.group(1)
        row_num = int(m.group(2))
        lag = float(m.group(3)) if m.group(3) else 0.0
        dep_type = "SS" if col == "E" else "FS"
        terms.append((dep_type, row_num, lag))
    return terms


def parse_relative_formula_excel(
    file_obj: BytesIO,
    project_id: int,
    sheet: int | str = 0,
) -> dict:
    """
    Parse formula-driven schedule workbook and return task rows + dependencies.

    The model assumes planned start/finish formulas in columns E/G and planned duration in F,
    with all dates relative to the first schedulable task start (anchor offset 0).
    """
    wb_formula = load_workbook(file_obj, data_only=False)
    file_obj.seek(0)
    wb_value = load_workbook(file_obj, data_only=True)

    ws_formula = wb_formula[sheet] if isinstance(sheet, str) else wb_formula.worksheets[sheet]
    ws_value = wb_value[ws_formula.title]

    header_row = None
    for r in range(1, min(40, ws_formula.max_row) + 1):
        vals = [str(ws_formula.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 20)]
        line = " ".join(v for v in vals if v)
        if "activity" in line and "planned" in line:
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not locate formula schedule header row in workbook.")

    col_idx: dict[str, int] = {}
    for c in range(1, ws_formula.max_column + 1):
        raw = str(ws_formula.cell(row=header_row, column=c).value or "").strip().lower()
        key = raw.replace(" ", "_")
        if key in ("s.no.", "s.no", "s_no", "sr_no"):
            col_idx["row_num"] = c
        elif key in ("activity", "task_name", "task"):
            col_idx["task_name"] = c
        elif key in ("quantity", "qty"):
            col_idx["quantity"] = c
        elif key in ("unit", "uom"):
            col_idx["unit"] = c
        elif key in ("planned_start", "plan_start"):
            col_idx["start"] = c
        elif key in ("planned_duration", "plan_duration", "duration"):
            col_idx["duration"] = c
        elif key in ("planned_finish", "plan_finish", "finish"):
            col_idx["finish"] = c
        elif key in ("percent_complete", "%_complete", "pct_complete"):
            col_idx["pct_complete"] = c

    required = ("task_name", "start", "duration", "finish")
    if not all(k in col_idx for k in required):
        raise ValueError("Workbook is missing required columns for formula schedule import.")

    row_entries: list[dict] = []
    for r in range(header_row + 1, ws_formula.max_row + 1):
        task_name = str(ws_formula.cell(row=r, column=col_idx["task_name"]).value or "").strip()
        if not task_name:
            continue

        qty = _parse_float(ws_formula.cell(row=r, column=col_idx.get("quantity", 0)).value) if "quantity" in col_idx else None
        unit = str(ws_formula.cell(row=r, column=col_idx.get("unit", 0)).value or "").strip() if "unit" in col_idx else ""
        dur = _parse_duration(ws_formula.cell(row=r, column=col_idx["duration"]).value)

        start_cell_formula = ws_formula.cell(row=r, column=col_idx["start"]).value
        finish_cell_formula = ws_formula.cell(row=r, column=col_idx["finish"]).value
        start_formula = start_cell_formula if isinstance(start_cell_formula, str) and start_cell_formula.startswith("=") else None
        finish_formula = finish_cell_formula if isinstance(finish_cell_formula, str) and finish_cell_formula.startswith("=") else None

        start_val = ws_value.cell(row=r, column=col_idx["start"]).value
        finish_val = ws_value.cell(row=r, column=col_idx["finish"]).value

        pct = 0.0
        if "pct_complete" in col_idx:
            pct_raw = str(ws_formula.cell(row=r, column=col_idx["pct_complete"]).value or "0").strip().replace("%", "")
            try:
                pct = float(pct_raw) if pct_raw.lower() != "nan" else 0.0
            except ValueError:
                pct = 0.0

        row_num = r
        if "row_num" in col_idx:
            try:
                row_num = int(float(str(ws_formula.cell(row=r, column=col_idx["row_num"]).value).strip()))
            except Exception:
                row_num = r

        is_summary = 1 if (qty is None and dur in (None, 0.0) and not start_formula and not finish_formula) else 0

        row_entries.append(
            {
                "excel_row": r,
                "row_num": row_num,
                "task_name": task_name,
                "quantity": qty,
                "unit": unit if unit and unit.lower() != "nan" else None,
                "duration_days": dur,
                "pct_complete": pct,
                "is_summary": is_summary,
                "start_formula": start_formula,
                "finish_formula": finish_formula,
                "start_value": start_val,
                "finish_value": finish_val,
            }
        )

    # Determine anchor from first non-summary task with a computed start date.
    anchor_date = None
    anchor_excel_row = None
    for entry in row_entries:
        if entry["is_summary"] == 1:
            continue
        sv = entry["start_value"]
        if isinstance(sv, datetime):
            anchor_date = sv.date()
            anchor_excel_row = entry["excel_row"]
            break
        if isinstance(sv, date):
            anchor_date = sv
            anchor_excel_row = entry["excel_row"]
            break
    if anchor_date is None:
        anchor_date = date.today()
        anchor_excel_row = row_entries[0]["excel_row"] if row_entries else 1

    offsets: dict[tuple[str, int], float] = {}
    if anchor_excel_row is not None:
        offsets[("E", anchor_excel_row)] = 0.0

    # Seed direct date values where available.
    for entry in row_entries:
        r = entry["excel_row"]
        sv = entry["start_value"]
        fv = entry["finish_value"]
        if isinstance(sv, datetime):
            offsets[("E", r)] = float((sv.date() - anchor_date).days)
        elif isinstance(sv, date):
            offsets[("E", r)] = float((sv - anchor_date).days)
        if isinstance(fv, datetime):
            offsets[("G", r)] = float((fv.date() - anchor_date).days)
        elif isinstance(fv, date):
            offsets[("G", r)] = float((fv - anchor_date).days)

    def _ref_value(token: str) -> float | None:
        token = token.upper().replace("$", "")
        m = re.fullmatch(r"([A-Z]+)(\d+)", token)
        if not m:
            return None
        col = m.group(1)
        rnum = int(m.group(2))
        if col in ("E", "G"):
            return offsets.get((col, rnum))
        if col == "F":
            for e in row_entries:
                if e["excel_row"] == rnum:
                    return float(e["duration_days"] or 0.0)
            return 0.0
        if col == "L" and rnum == 4:
            return 0.0
        return None

    def _eval_formula(expr: str) -> float | None:
        norm = _normalize_excel_formula(expr)
        if not norm.startswith("="):
            return None
        rhs = norm[1:]
        refs = set(re.findall(r"[A-Z]+\d+", rhs))
        replaced = rhs
        for ref in sorted(refs, key=len, reverse=True):
            val = _ref_value(ref)
            if val is None:
                return None
            replaced = replaced.replace(ref, str(val))
        replaced = replaced.replace("MAX", "max")
        if re.search(r"[^0-9+\-().,max]", replaced):
            return None
        return _safe_eval_numeric(replaced)

    unresolved = True
    max_iters = max(5, len(row_entries) * 3)
    i = 0
    while unresolved and i < max_iters:
        i += 1
        unresolved = False
        changed = False
        for entry in row_entries:
            r = entry["excel_row"]
            if ("E", r) not in offsets:
                sf = entry["start_formula"]
                if sf:
                    v = _eval_formula(sf)
                    if v is not None:
                        offsets[("E", r)] = float(v)
                        changed = True
                    else:
                        unresolved = True
            if ("G", r) not in offsets:
                ff = entry["finish_formula"]
                if ff:
                    v = _eval_formula(ff)
                    if v is not None:
                        offsets[("G", r)] = float(v)
                        changed = True
                    else:
                        unresolved = True
                else:
                    s = offsets.get(("E", r))
                    d = float(entry["duration_days"] or 0.0)
                    if s is not None:
                        offsets[("G", r)] = s + d
                        changed = True
        if not changed:
            break

    tasks: list[dict] = []
    dependencies: list[dict] = []
    for entry in row_entries:
        r = entry["excel_row"]
        s_off = offsets.get(("E", r))
        f_off = offsets.get(("G", r))
        s_date = (anchor_date + timedelta(days=s_off)).isoformat() if s_off is not None else None
        f_date = (anchor_date + timedelta(days=f_off)).isoformat() if f_off is not None else None

        wbs = str(entry["row_num"])
        tasks.append(
            {
                "project_id": project_id,
                "row_num": entry["row_num"],
                "wbs": wbs,
                "task_name": entry["task_name"],
                "duration_days": entry["duration_days"],
                "start_date": s_date,
                "finish_date": f_date,
                "predecessors": "",
                "pct_complete": entry["pct_complete"],
                "wbs_level": 1,
                "is_summary": entry["is_summary"],
                "notes": "",
                "site_id": None,
                "template_id": None,
                "quantity": entry["quantity"],
                "unit": entry["unit"],
                "planned_start_formula": entry["start_formula"],
                "planned_finish_formula": entry["finish_formula"],
                "planned_start_offset_days": s_off,
                "planned_finish_offset_days": f_off,
            }
        )

        for expr in [entry["start_formula"], entry["finish_formula"]]:
            if not expr:
                continue
            for dep_type, from_row_num, lag_days in _extract_dependency_terms(expr):
                dependencies.append(
                    {
                        "project_id": project_id,
                        "site_id": None,
                        "to_row_num": entry["row_num"],
                        "from_row_num": from_row_num,
                        "dep_type": dep_type,
                        "lag_days": lag_days,
                        "source_formula": expr,
                    }
                )

    return {
        "tasks": tasks,
        "dependencies": dependencies,
        "anchor_start_date": anchor_date.isoformat(),
    }
